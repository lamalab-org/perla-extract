"""Create and read the deliberately small offline-review workbook.

The workbook is an alternate editor for scalar values, not a second scientific
schema. Stable record identities, paths, and the source revision remain read-only;
the importer compares them with a freshly generated contract before returning any
changes. This keeps Excel convenient without allowing row edits to change record
structure or bypass whole-study validation in :mod:`review_workbench.study_review`.
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Iterable
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

FORMAT_NAME = "perla-review-workbook"
FORMAT_VERSION = 4
FIXED_SHEETS = ("Instructions", "Record review")
RECORD_HEADERS = (
    "Review outcome",
    "Reviewer note",
    "Record kind",
    "Record label",
    "Linked at",
    "Family",
    "Individual device",
    "_record_collection",
    "_record_id",
)
FIELD_HEADERS = (
    "Record label",
    "Field group",
    "Field",
    "Extracted value",
    "Corrected value",
    "Corrected value type",
    "Reviewer note",
    "Evidence quote",
    "Evidence block",
    "Linked at",
    "Family",
    "Individual device",
    "Schema path",
    "_editable",
    "_record_collection",
    "_record_id",
    "Record kind",
)
RECORD_COLLECTION_COLUMN = RECORD_HEADERS.index("_record_collection")
RECORD_ID_COLUMN = RECORD_HEADERS.index("_record_id")
RECORD_OUTCOME_COLUMN = RECORD_HEADERS.index("Review outcome")
RECORD_NOTE_COLUMN = RECORD_HEADERS.index("Reviewer note")
FIELD_COLLECTION_COLUMN = FIELD_HEADERS.index("_record_collection")
FIELD_ID_COLUMN = FIELD_HEADERS.index("_record_id")
FIELD_PATH_COLUMN = FIELD_HEADERS.index("Schema path")
FIELD_CURRENT_COLUMN = FIELD_HEADERS.index("Extracted value")
FIELD_REVIEWED_COLUMN = FIELD_HEADERS.index("Corrected value")
FIELD_TYPE_COLUMN = FIELD_HEADERS.index("Corrected value type")
FIELD_EDITABLE_COLUMN = FIELD_HEADERS.index("_editable")
FIELD_NOTE_COLUMN = FIELD_HEADERS.index("Reviewer note")
FIELD_EVIDENCE_BLOCK_COLUMN = FIELD_HEADERS.index("Evidence block")
FIELD_EVIDENCE_QUOTE_COLUMN = FIELD_HEADERS.index("Evidence quote")
OUTCOME_TO_DECISION = {
    "All fields match source": "verified",
    "Cannot establish from source": "uncertain",
    "Correct fields": "needs_correction",
}
DECISION_TO_OUTCOME = {value: key for key, value in OUTCOME_TO_DECISION.items()}
NO_OUTCOME = "Not reviewed"
MAX_WORKBOOK_BYTES = 15 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
BRAND = "176B52"
BRAND_SOFT = "E2F2EC"
EDITABLE_FILL = "FFF3BF"
READ_ONLY_FILL = "E7E9E8"
CHANGED_FILL = "FFE0B2"


@dataclass(frozen=True)
class WorkbookChange:
    """One scalar correction recovered from a validated workbook row."""

    collection: str
    record_id: str
    path: str
    value: Any
    note: str
    evidence: tuple[dict[str, str], ...]


@dataclass(frozen=True)
class WorkbookDecision:
    """One complete-record judgment recovered from the record-review sheet."""

    collection: str
    record_id: str
    decision: str
    note: str


@dataclass(frozen=True)
class WorkbookComment:
    """Preserve an Excel cell comment with enough row context to interpret it."""

    sheet: str
    cell: str
    text: str
    author: str
    record_collection: str | None
    record_id: str | None
    schema_path: str | None


@dataclass(frozen=True)
class WorkbookReview:
    """Validated edits and metadata ready for one atomic review transition."""

    base_revision: int
    scope_device_id: str | None
    changes: tuple[WorkbookChange, ...]
    decisions: tuple[WorkbookDecision, ...]
    comments: tuple[WorkbookComment, ...]
    sha256: str


@dataclass(frozen=True)
class _FieldRow:
    values: tuple[Any, ...]
    current_value: Any
    editable: bool


@dataclass(frozen=True)
class _RecordContext:
    """Human-readable relationship context derived from explicit schema links."""

    record_label: str
    link_scope: str
    family: str
    device: str


def _pointer_part(value: object) -> str:
    return str(value).replace("~", "~0").replace("/", "~1")


def _value_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    return "text"


def _citation(value: Any, inherited: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(value, dict):
        return inherited
    evidence = value.get("evidence")
    return evidence[0] if isinstance(evidence, list) and evidence else inherited


def _scalar_rows(
    collection: str,
    record_id: str,
    record_index: int,
    value: Any,
    label: str,
    context: _RecordContext,
    path: tuple[str | int, ...] = (),
    inherited_citation: dict[str, Any] | None = None,
) -> Iterable[_FieldRow]:
    """Flatten schema leaves while retaining paths and their nearest evidence.

    Excel cells should contain one atomic value. Recursion is generic over nested
    Pydantic output, while ``evidence`` is represented in dedicated support columns
    instead of being exposed as editable scientific data.
    """

    citation = _citation(value, inherited_citation)
    if isinstance(value, list):
        for index, item in enumerate(value):
            yield from _scalar_rows(
                collection,
                record_id,
                record_index,
                item,
                label,
                context,
                (*path, index),
                citation,
            )
        return
    if isinstance(value, dict):
        for key, item in value.items():
            if key != "evidence":
                yield from _scalar_rows(
                    collection,
                    record_id,
                    record_index,
                    item,
                    label,
                    context,
                    (*path, key),
                    citation,
                )
        return

    field = str(path[-1])
    editable = not field.endswith("_id")
    json_path = "/" + "/".join(
        _pointer_part(part) for part in (collection, record_index, *path)
    )
    yield _FieldRow(
        values=(
            context.record_label,
            " › ".join(str(part) for part in path[:-1]),
            field,
            value,
            value,
            _value_type(value),
            "",
            str(citation.get("quote", "")) if citation else "",
            str(citation.get("block_id", "")) if citation else "",
            context.link_scope,
            context.family,
            context.device,
            json_path,
            "Yes" if editable else "No",
            collection,
            record_id,
            label,
        ),
        current_value=value,
        editable=editable,
    )


def _records_in_scope(
    truth: dict[str, Any],
    identifiers: dict[str, str],
    device_id: str | None,
) -> list[tuple[str, int, dict[str, Any]]]:
    """Select either the paper or one device plus the context needed to review it."""

    if device_id is None:
        return [
            (collection, index, record)
            for collection in identifiers
            for index, record in enumerate(truth.get(collection, []))
        ]
    devices = truth.get("individual_devices", [])
    device = next(
        (
            item
            for item in devices
            if str(item.get(identifiers["individual_devices"])) == device_id
        ),
        None,
    )
    if device is None:
        raise ValueError(f"unknown device {device_id}")
    family_id = device.get(identifiers["device_families"])
    selected: list[tuple[str, int, dict[str, Any]]] = []
    for collection in identifiers:
        for index, record in enumerate(truth.get(collection, [])):
            include = False
            if collection == "device_families":
                include = record.get(identifiers[collection]) == family_id
            elif collection == "individual_devices":
                include = record.get(identifiers[collection]) == device_id
            elif collection in {"performance_observations", "stability_tests"}:
                linked_device = record.get(identifiers["individual_devices"])
                linked_family = record.get(identifiers["device_families"])
                include = linked_device == device_id or (
                    not linked_device and linked_family == family_id
                )
            elif collection == "population_statistics":
                include = record.get(identifiers["device_families"]) == family_id
            if include:
                selected.append((collection, index, record))
    return selected


def _entity_display(record: dict[str, Any] | None, record_id: str) -> str:
    """Pair a readable entity label with its immutable identifier."""

    if not record_id:
        return ""
    record = record or {}
    label = str(record.get("label") or record.get("specimen_label") or "").strip()
    return f"{label} [{record_id}]" if label and label != record_id else record_id


def _record_context(
    truth: dict[str, Any],
    record: dict[str, Any],
    record_id: str,
    identifiers: dict[str, str],
) -> _RecordContext:
    """Resolve explicit family and device links without inventing membership.

    Population statistics normally point only to a device family. The resulting
    ``Device family only`` label is intentional: devices that happen to share that
    family are context, but are not assumed to be members of the reported sample.
    """

    family_identifier = identifiers["device_families"]
    device_identifier = identifiers["individual_devices"]
    families = {
        str(item[family_identifier]): item
        for item in truth.get("device_families", [])
    }
    devices = {
        str(item[device_identifier]): item
        for item in truth.get("individual_devices", [])
    }
    device_id = str(record.get(device_identifier) or "")
    family_id = str(record.get(family_identifier) or "")
    device = devices.get(device_id)
    if device is not None and not family_id:
        family_id = str(device.get(family_identifier) or "")
    if device_id:
        link_scope = "Individual device"
    elif family_id:
        link_scope = "Device family (no individual device link)"
    else:
        link_scope = "No explicit device or family link"
    return _RecordContext(
        record_label=str(
            record.get("label") or record.get("specimen_label") or record_id
        ),
        link_scope=link_scope,
        family=_entity_display(families.get(family_id), family_id),
        device=_entity_display(device, device_id),
    )


def _contract(
    truth: dict[str, Any],
    identifiers: dict[str, str],
    labels: dict[str, str],
    device_id: str | None,
) -> tuple[list[tuple[Any, ...]], list[_FieldRow]]:
    records, fields = [], []
    for collection, index, record in _records_in_scope(
        truth, identifiers, device_id
    ):
        record_id = str(record[identifiers[collection]])
        context = _record_context(truth, record, record_id, identifiers)
        records.append(
            (
                NO_OUTCOME,
                "",
                labels[collection],
                context.record_label,
                context.link_scope,
                context.family,
                context.device,
                collection,
                record_id,
            )
        )
        fields.extend(
            _scalar_rows(
                collection, record_id, index, record, labels[collection], context
            )
        )
    priority = {
        "device_families": 0,
        "population_statistics": 1,
        "individual_devices": 2,
        "performance_observations": 3,
        "stability_tests": 4,
    }
    records.sort(
        key=lambda row: (
            str(row[RECORD_HEADERS.index("Family")]),
            str(row[RECORD_HEADERS.index("Individual device")]),
            priority.get(str(row[RECORD_COLLECTION_COLUMN]), 99),
            str(row[RECORD_HEADERS.index("Record label")]),
        )
    )
    return records, fields


def _style_sheet(sheet: Any, widths: tuple[int, ...]) -> None:
    """Apply shared presentation without defining sheet-level data behavior.

    Review grids are Excel tables, and each table owns its AutoFilter. Adding a
    second worksheet AutoFilter over the same range makes Excel repair the file
    by removing the table, even though permissive readers accept the package.
    """

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _hide_columns(sheet: Any, *headers: str) -> None:
    """Keep importer metadata in the workbook without burdening reviewers."""

    for header in headers:
        index = next(
            index for index, cell in enumerate(sheet[1], 1) if cell.value == header
        )
        sheet.column_dimensions[get_column_letter(index)].hidden = True


def _field_sheet_name(collection: str) -> str:
    """Turn a schema collection name into a stable, readable Excel tab name.

    Collection keys are the durable grouping already defined by the scientific
    schema. Reusing them avoids maintaining a second list of domain categories in
    the spreadsheet layer. Excel limits worksheet titles to 31 characters.
    """

    return collection.replace("_", " ").title()[:31]


def _field_rows_by_collection(
    field_rows: Iterable[_FieldRow], identifiers: dict[str, str]
) -> dict[str, list[_FieldRow]]:
    """Group flattened scalar rows by their existing top-level record collection."""

    grouped: dict[str, list[_FieldRow]] = {collection: [] for collection in identifiers}
    for row in field_rows:
        grouped[str(row.values[FIELD_COLLECTION_COLUMN])].append(row)
    return {collection: rows for collection, rows in grouped.items() if rows}


def _add_field_sheet(
    book: Workbook, collection: str, field_rows: list[_FieldRow]
) -> None:
    """Add one filterable correction tab for a scientific record collection."""

    sheet = book.create_sheet(_field_sheet_name(collection))
    sheet.append(FIELD_HEADERS)
    for field_row in field_rows:
        sheet.append(field_row.values)
    _keep_source_strings_literal(sheet)
    _style_sheet(
        sheet,
        (42, 34, 24, 24, 24, 19, 38, 68, 28, 34, 42, 42, 52, 11, 27, 32, 20),
    )
    sheet.freeze_panes = "D2"
    sheet.sheet_properties.tabColor = "5F8A7B"
    _hide_columns(
        sheet,
        "_editable",
        "_record_collection",
        "_record_id",
        "Record kind",
    )
    last_column = get_column_letter(len(FIELD_HEADERS))
    sheet.add_table(
        Table(
            displayName=f"Fields_{collection}",
            ref=f"A1:{last_column}{len(field_rows) + 1}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    type_validation = DataValidation(
        type="list", formula1='"text,integer,number,boolean,null"'
    )
    sheet.add_data_validation(type_validation)
    type_column = get_column_letter(FIELD_TYPE_COLUMN + 1)
    type_validation.add(f"{type_column}2:{type_column}{len(field_rows) + 1}")
    for row_index, field_row in enumerate(field_rows, 2):
        fill = EDITABLE_FILL if field_row.editable else READ_ONLY_FILL
        for column in (
            FIELD_REVIEWED_COLUMN,
            FIELD_TYPE_COLUMN,
            FIELD_NOTE_COLUMN,
            FIELD_EVIDENCE_BLOCK_COLUMN,
            FIELD_EVIDENCE_QUOTE_COLUMN,
        ):
            sheet.cell(row_index, column + 1).fill = PatternFill(
                "solid", fgColor=fill
            )
        if not field_row.editable:
            sheet.row_dimensions[row_index].hidden = True
    current_column = get_column_letter(FIELD_CURRENT_COLUMN + 1)
    reviewed_column = get_column_letter(FIELD_REVIEWED_COLUMN + 1)
    sheet.conditional_formatting.add(
        f"A2:{last_column}{len(field_rows) + 1}",
        FormulaRule(
            formula=[f"${current_column}2<>${reviewed_column}2"],
            fill=PatternFill("solid", fgColor=CHANGED_FILL),
        ),
    )


def _keep_source_strings_literal(sheet: Any) -> None:
    """Prevent source text beginning with ``=`` from becoming an Excel formula."""

    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"


def create_review_workbook(
    *,
    truth: dict[str, Any],
    identifiers: dict[str, str],
    labels: dict[str, str],
    paper_id: str,
    split: str,
    revision: int,
    schema_sha256: str,
    current_decisions: dict[str, str] | None = None,
    device_id: str | None = None,
) -> bytes:
    """Return an XLSX review form bound to one immutable study revision."""

    record_rows, field_rows = _contract(truth, identifiers, labels, device_id)
    if not record_rows:
        raise ValueError(
            "this paper has no records to review in Excel; add missing records in the browser first"
        )
    current_decisions = current_decisions or {}
    book = Workbook()
    instructions = book.active
    instructions.title = "Instructions"
    records = book.create_sheet("Record review")

    instructions.sheet_view.showGridLines = False
    instructions.sheet_properties.tabColor = BRAND
    instructions.merge_cells("A1:B1")
    instructions["A1"] = "PERLA offline review workbook"
    instructions["A1"].fill = PatternFill("solid", fgColor=BRAND)
    instructions["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 28
    guidance = (
        ("Paper", truth.get("paper", {}).get("title") or paper_id),
        ("Scope", f"Device {device_id} plus linked context" if device_id else "All paper records"),
        ("Start here", "Use Record review as the checklist. Choose one outcome for each complete record; you do not need to review every row on the detail tabs."),
        ("If correct", "Choose All fields match source. No detail-sheet edits are needed."),
        ("If uncertain", "Choose Cannot establish from source and explain the uncertainty in Reviewer note."),
        ("If wrong", "Choose Correct fields, then open that record type's tab and edit only the wrong scalar rows."),
        ("Corrections", "Yellow cells accept input. Keep each value atomic and provide a short note plus the exact supporting evidence quote and block."),
        ("Relationships", "Linked at, Family, and Individual device are explicit schema links. Family-only statistics are not silently assigned to one of the family's devices."),
        ("Structure", "Do not add, delete, or rename rows or sheets. Add, remove, or relink complete records in the browser."),
        ("Finish", "Upload this file to the same paper. The app validates all changes together and rejects stale or structurally altered workbooks."),
    )
    for row in guidance:
        instructions.append(row)
    instructions.append(("", ""))
    instructions.append(
        ("Workbook map", "Open a tab below only when you need its records or scalar values.")
    )
    instructions.append(("Record review", f"Main checklist · {len(record_rows)} complete records"))
    for collection, rows in _field_rows_by_collection(field_rows, identifiers).items():
        instructions.append(
            (
                _field_sheet_name(collection),
                f"Corrections for {labels[collection].lower()} records · "
                f"{sum(row.editable for row in rows)} editable scalar fields",
            )
        )
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 105
    separator = Side(style="thin", color="DCE3DF")
    for row in instructions.iter_rows(min_row=2):
        row[0].fill = PatternFill("solid", fgColor=BRAND_SOFT)
        row[0].font = Font(bold=True, color=BRAND)
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row:
            cell.border = Border(bottom=separator)
    for row_index in range(2, instructions.max_row + 1):
        instructions.row_dimensions[row_index].height = 34
    instructions.freeze_panes = "A2"

    records.append(RECORD_HEADERS)
    for row in record_rows:
        key = f"{row[RECORD_COLLECTION_COLUMN]}:{row[RECORD_ID_COLUMN]}"
        outcome = DECISION_TO_OUTCOME.get(current_decisions.get(key, ""), NO_OUTCOME)
        values = list(row)
        values[RECORD_OUTCOME_COLUMN] = outcome
        records.append(values)
    _keep_source_strings_literal(records)
    _style_sheet(records, (28, 44, 20, 48, 34, 46, 46, 27, 32))
    records.freeze_panes = "E2"
    records.sheet_properties.tabColor = BRAND
    _hide_columns(records, "_record_collection", "_record_id")
    record_last_column = get_column_letter(len(RECORD_HEADERS))
    records.add_table(
        Table(
            displayName="RecordReviewTable",
            ref=f"A1:{record_last_column}{len(record_rows) + 1}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
            ),
        )
    )
    outcome_validation = DataValidation(
        type="list",
        formula1='"Not reviewed,All fields match source,Cannot establish from source,Correct fields"',
    )
    records.add_data_validation(outcome_validation)
    outcome_column = get_column_letter(RECORD_OUTCOME_COLUMN + 1)
    outcome_validation.add(
        f"{outcome_column}2:{outcome_column}{len(record_rows) + 1}"
    )
    for row in records.iter_rows(
        min_row=2,
        min_col=RECORD_OUTCOME_COLUMN + 1,
        max_col=RECORD_NOTE_COLUMN + 1,
    ):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=EDITABLE_FILL)
    record_range = f"A2:{record_last_column}{len(record_rows) + 1}"
    for outcome, fill in (
        ("All fields match source", "DDF2E8"),
        ("Cannot establish from source", EDITABLE_FILL),
        ("Correct fields", CHANGED_FILL),
    ):
        records.conditional_formatting.add(
            record_range,
            FormulaRule(
                formula=[f'$A2="{outcome}"'],
                fill=PatternFill("solid", fgColor=fill),
            ),
        )

    grouped_fields = _field_rows_by_collection(field_rows, identifiers)
    for collection, collection_rows in grouped_fields.items():
        _add_field_sheet(book, collection, collection_rows)

    meta = book.create_sheet("_meta")

    metadata = {
        "format": FORMAT_NAME,
        "format_version": FORMAT_VERSION,
        "paper_id": paper_id,
        "split": split,
        "base_revision": revision,
        "scope_device_id": device_id or "",
        "schema_sha256": schema_sha256,
        "ground_truth_sha256": _json_digest(truth),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "record_count": len(record_rows),
        "field_count": len(field_rows),
    }
    meta.append(("key", "value"))
    for key, value in metadata.items():
        meta.append((key, value))
    _style_sheet(meta, (34, 78))
    meta.sheet_state = "hidden"

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _json_digest(value: object) -> str:
    encoded = json.dumps(
        value, sort_keys=True, separators=(",", ":"), ensure_ascii=False
    ).encode()
    return hashlib.sha256(encoded).hexdigest()


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decode_value(value: Any, kind: str) -> Any:
    if kind == "null":
        if value not in (None, ""):
            raise ValueError("a null value must have an empty reviewed-value cell")
        return None
    if kind == "text":
        return "" if value is None else str(value)
    if kind == "boolean":
        if isinstance(value, bool):
            return value
        normalized = _cell_text(value).lower()
        if normalized in {"true", "yes", "1"}:
            return True
        if normalized in {"false", "no", "0"}:
            return False
        raise ValueError(f"{value!r} is not a boolean")
    if kind == "integer":
        number = float(value)
        if not number.is_integer():
            raise ValueError(f"{value!r} is not an integer")
        return int(number)
    if kind == "number":
        return float(value)
    raise ValueError(f"unknown value type {kind!r}")


def _metadata(sheet: Any) -> dict[str, Any]:
    return {
        str(key): value
        for key, value in sheet.iter_rows(min_row=2, max_col=2, values_only=True)
        if key is not None
    }


def _rows(sheet: Any, headers: tuple[str, ...]) -> list[tuple[Any, ...]]:
    if any(cell.data_type == "f" for row in sheet.iter_rows() for cell in row):
        raise ValueError(f"{sheet.title} contains a formula; enter literal values only")
    actual_headers = tuple(cell.value for cell in sheet[1])
    if actual_headers != headers:
        raise ValueError(f"{sheet.title} columns were changed")
    return [
        tuple(cell.value for cell in row)
        for row in sheet.iter_rows(min_row=2, max_col=len(headers))
        if any(cell.value is not None for cell in row)
    ]


def _load_review_workbook(data: bytes) -> Workbook:
    """Apply upload limits once before any semantic workbook inspection."""

    if not data or len(data) > MAX_WORKBOOK_BYTES:
        raise ValueError("review workbook must be a non-empty XLSX smaller than 15 MiB")
    try:
        with ZipFile(BytesIO(data)) as archive:
            if sum(item.file_size for item in archive.infolist()) > MAX_UNCOMPRESSED_BYTES:
                raise ValueError("review workbook expands beyond the 100 MiB safety limit")
    except BadZipFile as error:
        raise ValueError("review workbook is not a readable XLSX file") from error
    try:
        book = load_workbook(
            BytesIO(data), data_only=False, read_only=False, keep_links=False
        )
    except Exception as error:  # malformed OOXML can fail in several XML readers
        raise ValueError("review workbook is not a readable XLSX file") from error
    return book


def _workbook_comments(book: Workbook) -> tuple[WorkbookComment, ...]:
    """Extract comments generically while attaching record or field identity when present."""

    comments = []
    for sheet in book.worksheets:
        if sheet.title == "_meta":
            continue
        headers = {
            _cell_text(cell.value): cell.column
            for cell in sheet[1]
            if _cell_text(cell.value)
        }
        for row in sheet.iter_rows():
            record_collection = (
                _cell_text(
                    sheet.cell(row[0].row, headers["_record_collection"]).value
                )
                if "_record_collection" in headers and row[0].row > 1
                else ""
            )
            record_id = (
                _cell_text(sheet.cell(row[0].row, headers["_record_id"]).value)
                if "_record_id" in headers and row[0].row > 1
                else ""
            )
            schema_path = (
                _cell_text(sheet.cell(row[0].row, headers["Schema path"]).value)
                if "Schema path" in headers and row[0].row > 1
                else ""
            )
            for cell in row:
                if cell.comment is None or not cell.comment.text.strip():
                    continue
                comments.append(
                    WorkbookComment(
                        sheet=sheet.title,
                        cell=cell.coordinate,
                        text=cell.comment.text.strip()[:8000],
                        author=(cell.comment.author or "")[:200],
                        record_collection=record_collection or None,
                        record_id=record_id or None,
                        schema_path=schema_path or None,
                    )
                )
    return tuple(comments)


def read_review_workbook_comments(
    data: bytes, *, paper_id: str, split: str
) -> tuple[int, tuple[WorkbookComment, ...], str]:
    """Recover comments from a stale PERLA workbook without applying stale values."""

    book = _load_review_workbook(data)
    if "_meta" not in book.sheetnames:
        raise ValueError("review workbook metadata sheet is missing")
    metadata = _metadata(book["_meta"])
    for key, expected in {
        "format": FORMAT_NAME,
        "paper_id": paper_id,
        "split": split,
    }.items():
        if metadata.get(key) != expected:
            raise ValueError(f"review workbook metadata does not match {key}")
    comments = _workbook_comments(book)
    return int(metadata.get("base_revision", 0)), comments, hashlib.sha256(data).hexdigest()


def read_review_workbook(
    data: bytes,
    *,
    truth: dict[str, Any],
    identifiers: dict[str, str],
    labels: dict[str, str],
    paper_id: str,
    split: str,
    revision: int,
    schema_sha256: str,
) -> WorkbookReview:
    """Validate an uploaded workbook and return its intentional review content."""

    book = _load_review_workbook(data)
    if "_meta" not in book.sheetnames:
        raise ValueError("review workbook metadata sheet is missing")
    metadata = _metadata(book["_meta"])
    expected_meta = {
        "format": FORMAT_NAME,
        "format_version": FORMAT_VERSION,
        "paper_id": paper_id,
        "split": split,
        "base_revision": revision,
        "schema_sha256": schema_sha256,
        "ground_truth_sha256": _json_digest(truth),
    }
    for key, expected in expected_meta.items():
        if metadata.get(key) != expected:
            if key in {"base_revision", "ground_truth_sha256"}:
                raise ValueError(
                    "this workbook was downloaded from an older paper revision; "
                    "download a new workbook and transfer the reviewed changes"
                )
            if key == "format_version":
                raise ValueError(
                    "this workbook uses an older layout; download a fresh workbook"
                )
            raise ValueError(f"review workbook metadata does not match {key}")
    device_id = _cell_text(metadata.get("scope_device_id")) or None
    expected_records, expected_fields = _contract(
        truth, identifiers, labels, device_id
    )
    grouped_fields = _field_rows_by_collection(expected_fields, identifiers)
    expected_sheets = (
        *FIXED_SHEETS,
        *(_field_sheet_name(collection) for collection in grouped_fields),
        "_meta",
    )
    if tuple(book.sheetnames) != expected_sheets:
        raise ValueError("review workbook sheets were added, removed, or renamed")
    record_rows = _rows(book["Record review"], RECORD_HEADERS)
    field_rows = [
        (sheet_name, row_number, row)
        for collection in grouped_fields
        for sheet_name in (_field_sheet_name(collection),)
        for row_number, row in enumerate(
            _rows(book[sheet_name], FIELD_HEADERS), start=2
        )
    ]
    if len(record_rows) != len(expected_records) or len(field_rows) != len(expected_fields):
        raise ValueError("review workbook rows were added, removed, or left incomplete")

    expected_record_map = {
        (row[RECORD_COLLECTION_COLUMN], row[RECORD_ID_COLUMN]): row
        for row in expected_records
    }
    actual_record_keys = [
        (
            _cell_text(row[RECORD_COLLECTION_COLUMN]),
            _cell_text(row[RECORD_ID_COLUMN]),
        )
        for row in record_rows
    ]
    if len(set(actual_record_keys)) != len(actual_record_keys) or set(
        actual_record_keys
    ) != set(expected_record_map):
        raise ValueError("Record review identities were added, removed, or duplicated")
    decisions: list[WorkbookDecision] = []
    for row_number, actual in enumerate(record_rows, 2):
        expected = expected_record_map[
            (
                _cell_text(actual[RECORD_COLLECTION_COLUMN]),
                _cell_text(actual[RECORD_ID_COLUMN]),
            )
        ]
        immutable_columns = tuple(
            index
            for index in range(len(RECORD_HEADERS))
            if index not in {RECORD_OUTCOME_COLUMN, RECORD_NOTE_COLUMN}
        )
        if any(
            _cell_text(actual[index]) != _cell_text(expected[index])
            for index in immutable_columns
        ):
            raise ValueError(f"Record review row {row_number} identity was changed")
        outcome = _cell_text(actual[RECORD_OUTCOME_COLUMN]) or NO_OUTCOME
        if outcome not in {NO_OUTCOME, *OUTCOME_TO_DECISION}:
            raise ValueError(f"Record review row {row_number} has an unknown outcome")
        if outcome != NO_OUTCOME:
            decisions.append(
                WorkbookDecision(
                    collection=str(actual[RECORD_COLLECTION_COLUMN]),
                    record_id=str(actual[RECORD_ID_COLUMN]),
                    decision=OUTCOME_TO_DECISION[outcome],
                    note=_cell_text(actual[RECORD_NOTE_COLUMN]),
                )
            )

    expected_field_map = {
        str(row.values[FIELD_PATH_COLUMN]): row for row in expected_fields
    }
    actual_field_keys = [
        _cell_text(row[FIELD_PATH_COLUMN]) for _, _, row in field_rows
    ]
    if len(set(actual_field_keys)) != len(actual_field_keys) or set(
        actual_field_keys
    ) != set(expected_field_map):
        raise ValueError("Field correction paths were added, removed, or duplicated")
    changes: list[WorkbookChange] = []
    for sheet_name, row_number, actual in field_rows:
        expected = expected_field_map[_cell_text(actual[FIELD_PATH_COLUMN])]
        expected_values = expected.values
        mutable_columns = {
            FIELD_REVIEWED_COLUMN,
            FIELD_TYPE_COLUMN,
            FIELD_NOTE_COLUMN,
            FIELD_EVIDENCE_BLOCK_COLUMN,
            FIELD_EVIDENCE_QUOTE_COLUMN,
        }
        immutable_text_columns = tuple(
            index
            for index in range(len(FIELD_HEADERS))
            if index not in {*mutable_columns, FIELD_CURRENT_COLUMN}
        )
        if any(
            _cell_text(actual[index]) != _cell_text(expected_values[index])
            for index in immutable_text_columns
        ):
            raise ValueError(f"{sheet_name} row {row_number} identity was changed")
        try:
            current = _decode_value(
                actual[FIELD_CURRENT_COLUMN], _value_type(expected.current_value)
            )
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"{sheet_name} row {row_number} current value was changed"
            ) from error
        if current != expected.current_value:
            raise ValueError(f"{sheet_name} row {row_number} current value was changed")
        reviewed_type = _cell_text(actual[FIELD_TYPE_COLUMN])
        try:
            reviewed = _decode_value(actual[FIELD_REVIEWED_COLUMN], reviewed_type)
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"{sheet_name} row {row_number} has an invalid reviewed value: {error}"
            ) from error
        if reviewed == expected.current_value and reviewed_type == _value_type(
            expected.current_value
        ):
            continue
        if not expected.editable:
            raise ValueError(f"{sheet_name} row {row_number} is read-only")
        note = _cell_text(actual[FIELD_NOTE_COLUMN])
        block_id = _cell_text(actual[FIELD_EVIDENCE_BLOCK_COLUMN])
        quote = _cell_text(actual[FIELD_EVIDENCE_QUOTE_COLUMN])
        if not note:
            raise ValueError(f"{sheet_name} row {row_number} needs a reviewer note")
        if not block_id or not quote:
            raise ValueError(f"{sheet_name} row {row_number} needs exact evidence")
        changes.append(
            WorkbookChange(
                collection=str(actual[FIELD_COLLECTION_COLUMN]),
                record_id=str(actual[FIELD_ID_COLUMN]),
                path=str(actual[FIELD_PATH_COLUMN]),
                value=reviewed,
                note=note,
                evidence=({"block_id": block_id, "quote": quote},),
            )
        )
    comments = _workbook_comments(book)
    if not changes and not decisions and not comments:
        raise ValueError(
            "review workbook contains no decisions, corrections, or cell comments"
        )
    return WorkbookReview(
        base_revision=revision,
        scope_device_id=device_id,
        changes=tuple(changes),
        decisions=tuple(decisions),
        comments=comments,
        sha256=hashlib.sha256(data).hexdigest(),
    )
