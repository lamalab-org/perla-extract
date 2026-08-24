from __future__ import annotations

import copy
import json
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

import pytest
from openpyxl import load_workbook

from perla_extract.study_extraction.models import (
    STUDY_SCHEMA_VERSION,
    study_schema_sha256,
)
from review_workbench.study_review import (
    InventoryAuditRequest,
    MainTextFigureCensus,
    MutationRequest,
    RecordDecisionRequest,
    StageRequest,
    StudyReviewStore,
    UndoMutationRequest,
)


def study_with_family(empty_study):
    study = copy.deepcopy(empty_study)
    study["device_families"] = [
        {
            "family_id": "family-control",
            "label": "Control",
            "variant": None,
            "architecture": None,
            "polarity": "not_reported",
            "full_stack_raw": None,
            "layers": [],
            "absorber_formula": None,
            "absorber_properties": [],
            "absorber_constituents": [],
            "processing_steps": [],
            "evidence": [
                {
                    "block_id": "main_p1_text_1",
                    "quote": "champion device",
                }
            ],
        }
    ]
    return study


def seed(store, empty_study, document_payload):
    return store.import_seed(
        "calibration",
        "10.0000--example",
        empty_study,
        document=document_payload,
        manifest={"model": "frontier"},
        reviewer_id="ada",
    )


def test_seed_is_immutable_and_truth_is_versioned(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    bundle = seed(store, empty_study, document_payload)
    assert bundle["revision"] == 1

    request = MutationRequest.model_validate(
        {
            "action": "replace",
            "path": "/unresolved_notes/0",
            "value": "Checked against the paper",
            "evidence": [{"block_id": "main_p1_text_1", "quote": "champion device"}],
            "note": "Resolve the model note",
            "base_revision": 1,
        }
    )
    updated = store.mutate("calibration", "10.0000--example", request, "ada")

    assert updated["ground_truth"]["unresolved_notes"] == ["Checked against the paper"]
    assert json.loads(store.seed_path("calibration", "10.0000--example").read_text())[
        "unresolved_notes"
    ] == ["Initial model note"]
    assert updated["events"][-1]["before"] == "Initial model note"
    assert updated["revision"] == 2
    assert bundle["manifest"]["schema_version"] == STUDY_SCHEMA_VERSION
    assert bundle["manifest"]["schema_sha256"] == study_schema_sha256()
    assert bundle["schema_compatibility"]["exact_match"] is True


def test_reviewer_can_undo_an_untouched_saved_correction(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    corrected = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="replace",
            path="/unresolved_notes/0",
            value="Checked against the paper",
            evidence=[{"block_id": "main_p1_text_1", "quote": "champion device"}],
            base_revision=1,
        ),
        "ada",
    )
    correction_id = corrected["events"][-1]["event_id"]

    progress = store.reviewer_progress("calibration", "ada")["papers"][0]
    assert progress["undoable_event_ids"] == [correction_id]

    with pytest.raises(PermissionError, match="only your own"):
        store.undo_mutation(
            "calibration",
            "10.0000--example",
            UndoMutationRequest(event_id=correction_id, base_revision=2),
            "grace",
        )

    undone = store.undo_mutation(
        "calibration",
        "10.0000--example",
        UndoMutationRequest(event_id=correction_id, base_revision=2),
        "ada",
    )

    assert undone["ground_truth"]["unresolved_notes"] == ["Initial model note"]
    assert undone["events"][-1]["details"] == {"undoes_event_id": correction_id}
    assert undone["events"][-1]["before"] == "Checked against the paper"
    assert undone["events"][-1]["after"] == "Initial model note"
    progress = store.reviewer_progress("calibration", "ada")["papers"][0]
    assert progress["undoable_event_ids"] == []
    assert progress["undone_event_ids"] == [correction_id]

    with pytest.raises(ValueError, match="already been undone"):
        store.undo_mutation(
            "calibration",
            "10.0000--example",
            UndoMutationRequest(event_id=correction_id, base_revision=3),
            "ada",
        )


def test_review_workbook_import_is_atomic_attributable_and_undoable(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    seed(store, study_with_family(empty_study), document_payload)
    data = store.review_workbook("calibration", "10.0000--example", "ada")
    book = load_workbook(BytesIO(data))
    assert book.sheetnames == [
        "Instructions",
        "Record review",
        "Device Families",
        "_meta",
    ]
    book["Record review"]["G2"] = "All fields match source"
    fields = book["Device Families"]
    label_row = next(
        row
        for row in range(2, fields.max_row + 1)
        if fields.cell(row, 5).value == "label"
    )
    fields.cell(label_row, 8).value = "Reviewed control"
    fields.cell(label_row, 11).value = "Corrected the family label."
    output = BytesIO()
    book.save(output)

    updated = store.import_review_workbook(
        "calibration",
        "10.0000--example",
        output.getvalue(),
        "ada",
        filename="family-review.xlsx",
    )

    assert updated["revision"] == 2
    assert updated["ground_truth"]["device_families"][0]["label"] == (
        "Reviewed control"
    )
    event = updated["events"][-1]
    assert event["kind"] == "spreadsheet_review"
    assert event["reviewer_id"] == "ada"
    assert event["details"]["filename"] == "family-review.xlsx"
    assert event["details"]["changed_fields"] == [
        {
            "path": "/device_families/0/label",
            "note": "Corrected the family label.",
        }
    ]
    assert updated["summary"]["record_decisions"]["ada"] == {
        "device_families:family-control": "verified"
    }

    progress = store.reviewer_progress("calibration", "ada")["papers"][0]
    assert progress["undoable_event_ids"] == [event["event_id"]]
    undone = store.undo_mutation(
        "calibration",
        "10.0000--example",
        UndoMutationRequest(event_id=event["event_id"], base_revision=2),
        "ada",
    )
    assert undone["ground_truth"]["device_families"][0]["label"] == "Control"
    assert undone["summary"]["record_decisions"].get("ada", {}) == {}


def test_review_workbook_groups_fields_by_scientific_record_type(
    tmp_path, empty_study, document_payload
):
    study = study_with_family(empty_study)
    citation = {"block_id": "main_p1_text_1", "quote": "champion device"}
    study["individual_devices"] = [
        {
            "device_id": "device-1",
            "family_id": "family-control",
            "label": "Champion device",
            "variant": None,
            "champion_status": "yes",
            "selection_basis": "champion",
            "reported_properties": [],
            "evidence": [citation],
        }
    ]
    study["performance_observations"] = [
        {
            "observation_id": "observation-1",
            "device_id": "device-1",
            "measurement_type": "jv_scan",
            "scan_direction": "not_reported",
            "metrics": [
                {
                    "name": "PCE",
                    "raw_value": "24.1%",
                    "value_number": 24.1,
                    "unit": "%",
                    "evidence": [citation],
                }
            ],
            "evidence": [citation],
        }
    ]
    store = StudyReviewStore(tmp_path)
    seed(store, study, document_payload)

    book = load_workbook(
        BytesIO(store.review_workbook("calibration", "10.0000--example", "ada"))
    )

    assert book.sheetnames == [
        "Instructions",
        "Record review",
        "Device Families",
        "Individual Devices",
        "Performance Observations",
        "_meta",
    ]
    assert {row[1].value for row in book["Device Families"].iter_rows(min_row=2)} == {
        "device_families"
    }
    assert {
        row[1].value for row in book["Individual Devices"].iter_rows(min_row=2)
    } == {"individual_devices"}
    assert {
        row[1].value
        for row in book["Performance Observations"].iter_rows(min_row=2)
    } == {"performance_observations"}


def test_review_workbook_rejects_stale_or_structurally_changed_files(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    seed(store, study_with_family(empty_study), document_payload)
    original = store.review_workbook("calibration", "10.0000--example", "ada")

    altered_book = load_workbook(BytesIO(original))
    altered_book["Record review"].delete_rows(2)
    altered = BytesIO()
    altered_book.save(altered)
    with pytest.raises(ValueError, match="added, removed"):
        store.import_review_workbook(
            "calibration", "10.0000--example", altered.getvalue(), "ada"
        )

    legacy_book = load_workbook(BytesIO(original))
    format_row = next(
        row
        for row in range(2, legacy_book["_meta"].max_row + 1)
        if legacy_book["_meta"].cell(row, 1).value == "format_version"
    )
    legacy_book["_meta"].cell(format_row, 2).value = 1
    legacy = BytesIO()
    legacy_book.save(legacy)
    with pytest.raises(ValueError, match="older layout"):
        store.import_review_workbook(
            "calibration", "10.0000--example", legacy.getvalue(), "ada"
        )

    store.decide_record(
        "calibration",
        "10.0000--example",
        RecordDecisionRequest(
            collection="device_families",
            record_id="family-control",
            decision="verified",
            base_revision=1,
        ),
        "ada",
    )
    with pytest.raises(ValueError, match="older paper revision"):
        store.import_review_workbook(
            "calibration", "10.0000--example", original, "ada"
        )


def test_undo_finds_an_appended_value_but_never_overwrites_later_work(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    added = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="add",
            path="/unresolved_notes/-",
            value="Temporary reviewer note",
            evidence=[{"block_id": "main_p1_text_1", "quote": "champion device"}],
            base_revision=1,
        ),
        "ada",
    )
    added_id = added["events"][-1]["event_id"]
    restored = store.undo_mutation(
        "calibration",
        "10.0000--example",
        UndoMutationRequest(event_id=added_id, base_revision=2),
        "ada",
    )
    assert restored["ground_truth"]["unresolved_notes"] == ["Initial model note"]
    assert restored["events"][-1]["path"] == "/unresolved_notes/1"

    corrected = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="replace",
            path="/unresolved_notes/0",
            value="Ada's correction",
            evidence=[{"block_id": "main_p1_text_1", "quote": "champion device"}],
            base_revision=restored["revision"],
        ),
        "ada",
    )
    correction_id = corrected["events"][-1]["event_id"]
    later = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="replace",
            path="/unresolved_notes/0",
            value="Grace's later correction",
            evidence=[{"block_id": "main_p1_text_1", "quote": "champion device"}],
            base_revision=corrected["revision"],
        ),
        "grace",
    )

    progress = store.reviewer_progress("calibration", "ada")["papers"][0]
    assert correction_id not in progress["undoable_event_ids"]
    with pytest.raises(ValueError, match="has since changed"):
        store.undo_mutation(
            "calibration",
            "10.0000--example",
            UndoMutationRequest(
                event_id=correction_id, base_revision=later["revision"]
            ),
            "ada",
        )


def test_undo_can_restore_a_removed_value(tmp_path, empty_study, document_payload):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    removed = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="remove",
            path="/unresolved_notes/0",
            note="Removed this note by mistake",
            base_revision=1,
        ),
        "ada",
    )

    restored = store.undo_mutation(
        "calibration",
        "10.0000--example",
        UndoMutationRequest(
            event_id=removed["events"][-1]["event_id"], base_revision=2
        ),
        "ada",
    )

    assert restored["ground_truth"]["unresolved_notes"] == ["Initial model note"]
    assert restored["events"][-1]["action"] == "add"


def test_reviewer_progress_contains_only_that_reviewers_persisted_work(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    store.import_seed(
        "calibration",
        "10.0000--example",
        empty_study,
        document=document_payload,
        manifest={},
        reviewer_id="seed-import",
    )
    audited = store.inventory_audit(
        "calibration",
        "10.0000--example",
        InventoryAuditRequest(
            base_revision=1,
            searched_sources=["main"],
            expected_counts={"individual_devices": 0},
            main_text_figure_census=MainTextFigureCensus(
                figures_reviewed=4,
                schema_relevant_figures=2,
                figure_only_records=1,
                figure_only_atomic_values=3,
                notes="Figure 3 contains stability values absent from the caption.",
            ),
            missing_or_ambiguous="No missing devices found.",
        ),
        "ada",
    )
    corrected = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="replace",
            path="/unresolved_notes/0",
            value="Ada checked this note",
            evidence=[{"block_id": "main_p1_text_1", "quote": "champion device"}],
            note="Checked against the source",
            base_revision=audited["revision"],
        ),
        "ada",
    )
    store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="replace",
            path="/unresolved_notes/0",
            value="Grace checked this note",
            evidence=[{"block_id": "main_p1_text_1", "quote": "champion device"}],
            base_revision=corrected["revision"],
        ),
        "grace",
    )

    progress = store.reviewer_progress("calibration", "ada")

    assert (progress["paper_count"], progress["annotation_count"]) == (1, 2)
    paper = progress["papers"][0]
    assert paper["paper_id"] == "10.0000--example"
    assert [event["kind"] for event in paper["events"]] == [
        "inventory_audit",
        "mutation",
    ]
    assert paper["events"][1]["before"] == "Initial model note"
    assert paper["events"][1]["after"] == "Ada checked this note"
    assert paper["current_inventory_audit"]["missing_or_ambiguous"] == (
        "No missing devices found."
    )
    assert paper["current_inventory_audit"]["review_scope_sources"] == ["main"]
    assert "searched_sources" not in paper["current_inventory_audit"]
    assert paper["current_inventory_audit"]["main_text_figure_census"] == {
        "figures_reviewed": 4,
        "schema_relevant_figures": 2,
        "figure_only_records": 1,
        "figure_only_atomic_values": 3,
        "notes": "Figure 3 contains stability values absent from the caption.",
    }
    assert all(event["reviewer_id"] == "ada" for event in paper["events"])
    assert store.reviewer_progress("calibration", "nobody")["papers"] == []


def test_main_text_figure_census_rejects_incoherent_counts():
    with pytest.raises(ValueError, match="cannot exceed figures reviewed"):
        MainTextFigureCensus(figures_reviewed=1, schema_relevant_figures=2)

    with pytest.raises(ValueError, match="require a schema-relevant figure"):
        MainTextFigureCensus(
            figures_reviewed=3,
            schema_relevant_figures=0,
            figure_only_atomic_values=1,
        )


def test_summary_renames_legacy_searched_sources_without_rewriting_the_event(
    empty_study,
):
    event = {
        "kind": "inventory_audit",
        "reviewer_id": "ada",
        "details": {"searched_sources": ["main"], "expected_counts": {}},
    }

    summary = StudyReviewStore.summary(empty_study, [event])

    assert summary["inventory_audits"]["ada"]["review_scope_sources"] == ["main"]
    assert "searched_sources" not in summary["inventory_audits"]["ada"]
    assert event["details"]["searched_sources"] == ["main"]


def test_bundle_marks_readable_older_schema_as_not_exactly_comparable(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    bundle = seed(store, empty_study, document_payload)
    source = store.storage.load_source("calibration", "10.0000--example")
    source.manifest["schema_version"] = 1
    source.manifest["schema_sha256"] = "historical"
    source_path = store.root / "state" / "sources" / "calibration" / "10.0000--example.json"
    payload = json.loads(source_path.read_text())
    payload["manifest"] = source.manifest
    source_path.write_text(json.dumps(payload))

    reloaded = store.load_bundle("calibration", "10.0000--example")

    assert bundle["schema_compatibility"]["exact_match"] is True
    assert reloaded["schema_compatibility"] == {
        "seed_schema_version": 1,
        "current_schema_version": STUDY_SCHEMA_VERSION,
        "seed_schema_sha256": "historical",
        "current_schema_sha256": study_schema_sha256(),
        "exact_match": False,
        "readable_by_current_schema": True,
    }


def test_human_citations_use_the_extraction_evidence_policy(
    tmp_path, empty_study, document_payload
):
    """Review should tolerate harmless OCR spacing exactly as extraction does."""

    document_payload["blocks"][0]["text"] = (
        "The absorber was Cs0.3FA0.6DMA0.1Pb (I 0.7 Br0.3)3."
    )
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    result = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="replace",
            path="/unresolved_notes/0",
            value="Formula checked",
            evidence=[
                {
                    "block_id": "main_p1_text_1",
                    "quote": "Cs0.3FA0.6DMA0.1Pb(I0.7Br0.3)3",
                }
            ],
            base_revision=1,
        ),
        "ada",
    )
    assert result["revision"] == 2


def test_corrections_require_supplied_exact_evidence(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    request = MutationRequest.model_validate(
        {
            "action": "replace",
            "path": "/unresolved_notes/0",
            "value": "corrected",
            "evidence": [{"block_id": "main_p1_text_1", "quote": "not in the block"}],
            "base_revision": 1,
        }
    )
    with pytest.raises(ValueError, match="quote is not present"):
        store.mutate("calibration", "10.0000--example", request, "ada")


def test_no_op_replacements_are_rejected(tmp_path, empty_study, document_payload):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    request = MutationRequest.model_validate(
        {
            "action": "replace",
            "path": "/unresolved_notes/0",
            "value": "Initial model note",
            "evidence": [{"block_id": "main_p1_text_1", "quote": "champion device"}],
            "base_revision": 1,
        }
    )
    with pytest.raises(ValueError, match="does not change"):
        store.mutate("calibration", "10.0000--example", request, "ada")


def test_referenced_records_must_be_reassigned_or_removed_before_their_target(
    tmp_path, empty_study, document_payload
):
    study = copy.deepcopy(empty_study)
    citation = {"block_id": "main_p1_text_1", "quote": "champion device"}
    study["individual_devices"] = [
        {
            "device_id": "device-1",
            "family_id": None,
            "label": "Champion device",
            "variant": None,
            "champion_status": "yes",
            "selection_basis": "champion",
            "reported_properties": [],
            "evidence": [citation],
        }
    ]
    study["performance_observations"] = [
        {
            "observation_id": "observation-1",
            "device_id": "device-1",
            "measurement_type": "jv_scan",
            "scan_direction": "not_reported",
            "metrics": [
                {
                    "name": "PCE",
                    "raw_value": "24.1%",
                    "value_number": 24.1,
                    "unit": "%",
                    "evidence": [citation],
                }
            ],
            "evidence": [citation],
        }
    ]
    store = StudyReviewStore(tmp_path)
    seed(store, study, document_payload)
    bundle = store.load_bundle("calibration", "10.0000--example")
    assert bundle["summary"]["record_references"] == {
        "individual_devices:device-1": [
            "performance_observations:observation-1"
        ]
    }

    remove_device = MutationRequest(
        action="remove",
        path="/individual_devices/0",
        note="This duplicate device is not supported.",
        base_revision=1,
    )
    with pytest.raises(ValueError, match="performance_observations:observation-1"):
        store.mutate("calibration", "10.0000--example", remove_device, "ada")

    without_observation = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="remove",
            path="/performance_observations/0",
            note="The observation belongs to the duplicate record.",
            base_revision=1,
        ),
        "ada",
    )
    removed = store.mutate(
        "calibration",
        "10.0000--example",
        remove_device.model_copy(update={"base_revision": without_observation["revision"]}),
        "ada",
    )
    assert removed["ground_truth"]["individual_devices"] == []


def test_blind_inventory_precedes_inventory_completion(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    with pytest.raises(ValueError, match="blind inventory"):
        store.complete_stage(
            "calibration",
            "10.0000--example",
            StageRequest(stage="inventory", base_revision=1, note=""),
            "ada",
        )

    audited = store.inventory_audit(
        "calibration",
        "10.0000--example",
        InventoryAuditRequest(
            base_revision=1,
            searched_sources=["main", "supplement"],
            expected_counts={"individual_devices": 2},
            missing_or_ambiguous="One stabilized specimen needs linkage.",
        ),
        "ada",
    )
    completed = store.complete_stage(
        "calibration",
        "10.0000--example",
        StageRequest(stage="inventory", base_revision=audited["revision"], note=""),
        "ada",
    )
    assert completed["summary"]["completed_stages"]["inventory"] == ["ada"]


def test_review_stages_cannot_be_skipped(tmp_path, empty_study, document_payload):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    with pytest.raises(ValueError, match="inventory stage"):
        store.complete_stage(
            "calibration",
            "10.0000--example",
            StageRequest(stage="fields", base_revision=1, note=""),
            "ada",
        )


def test_field_completion_requires_current_record_decisions(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    store.import_seed(
        "calibration",
        "10.0000--example",
        study_with_family(empty_study),
        document=document_payload,
        manifest={},
        reviewer_id="ada",
    )
    audited = store.inventory_audit(
        "calibration",
        "10.0000--example",
        InventoryAuditRequest(
            base_revision=1,
            searched_sources=["main"],
            expected_counts={"device_families": 1},
        ),
        "ada",
    )
    inventory = store.complete_stage(
        "calibration",
        "10.0000--example",
        StageRequest(stage="inventory", base_revision=audited["revision"]),
        "ada",
    )
    with pytest.raises(ValueError, match="1 remaining"):
        store.complete_stage(
            "calibration",
            "10.0000--example",
            StageRequest(stage="fields", base_revision=inventory["revision"]),
            "ada",
        )

    decided = store.decide_record(
        "calibration",
        "10.0000--example",
        RecordDecisionRequest(
            collection="device_families",
            record_id="family-control",
            decision="verified",
            base_revision=inventory["revision"],
        ),
        "ada",
    )
    completed = store.complete_stage(
        "calibration",
        "10.0000--example",
        StageRequest(stage="fields", base_revision=decided["revision"]),
        "ada",
    )
    assert completed["summary"]["completed_stages"]["fields"] == ["ada"]


def test_edit_invalidates_record_decision(tmp_path, empty_study, document_payload):
    store = StudyReviewStore(tmp_path)
    store.import_seed(
        "calibration",
        "10.0000--example",
        study_with_family(empty_study),
        document=document_payload,
        manifest={},
        reviewer_id="ada",
    )
    decided = store.decide_record(
        "calibration",
        "10.0000--example",
        RecordDecisionRequest(
            collection="device_families",
            record_id="family-control",
            decision="verified",
            base_revision=1,
        ),
        "ada",
    )
    changed = copy.deepcopy(decided["ground_truth"]["device_families"][0])
    changed["label"] = "Corrected control"
    corrected = store.mutate(
        "calibration",
        "10.0000--example",
        MutationRequest(
            action="replace",
            path="/device_families/0",
            value=changed,
            evidence=[{"block_id": "main_p1_text_1", "quote": "champion device"}],
            base_revision=decided["revision"],
        ),
        "ada",
    )
    assert corrected["summary"]["record_decisions"].get("ada", {}) == {}


def test_stage_cannot_be_completed_twice(tmp_path, empty_study, document_payload):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    audited = store.inventory_audit(
        "calibration",
        "10.0000--example",
        InventoryAuditRequest(
            base_revision=1, searched_sources=["main"], expected_counts={}
        ),
        "ada",
    )
    completed = store.complete_stage(
        "calibration",
        "10.0000--example",
        StageRequest(stage="inventory", base_revision=audited["revision"]),
        "ada",
    )
    with pytest.raises(ValueError, match="already complete"):
        store.complete_stage(
            "calibration",
            "10.0000--example",
            StageRequest(stage="inventory", base_revision=completed["revision"]),
            "ada",
        )


def test_stale_revisions_cannot_overwrite_other_reviewers(
    tmp_path, empty_study, document_payload
):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    store.inventory_audit(
        "calibration",
        "10.0000--example",
        InventoryAuditRequest(
            base_revision=1, searched_sources=["main"], expected_counts={}
        ),
        "ada",
    )
    request = MutationRequest.model_validate(
        {
            "action": "replace",
            "path": "/unresolved_notes/0",
            "value": "old write",
            "evidence": [{"block_id": "main_p1_text_1", "quote": "champion device"}],
            "base_revision": 1,
        }
    )
    with pytest.raises(ValueError, match="stale revision"):
        store.mutate("calibration", "10.0000--example", request, "grace")


def test_concurrent_reviews_create_exactly_one_next_revision(
    tmp_path, empty_study, document_payload
):
    """The immutable revision path is the winner election, not a process lock."""

    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)

    def audit(reviewer: str):
        try:
            store.inventory_audit(
                "calibration",
                "10.0000--example",
                InventoryAuditRequest(
                    base_revision=1,
                    searched_sources=["main"],
                    expected_counts={"individual_devices": 1},
                ),
                reviewer,
            )
            return "committed"
        except ValueError as error:
            assert "stale revision" in str(error)
            return "stale"

    with ThreadPoolExecutor(max_workers=2) as executor:
        outcomes = list(executor.map(audit, ["ada", "grace"]))

    assert sorted(outcomes) == ["committed", "stale"]
    final = store.load_bundle("calibration", "10.0000--example")
    assert final["revision"] == 2
    assert len(final["events"]) == 2


def test_lists_rich_record_counts(tmp_path, empty_study, document_payload):
    store = StudyReviewStore(tmp_path)
    seed(store, empty_study, document_payload)
    paper = store.list_papers("calibration")[0]
    assert paper["id"] == "10.0000--example"
    assert paper["device_families"] == 0
    assert paper["revision"] == 1
